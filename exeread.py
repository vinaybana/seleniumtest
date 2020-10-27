import csv
import os, shutil
from urllib.parse import urljoin, urlparse
import urllib.request
import pymysql, sqlparse
import openpyxl
from openpyxl import load_workbook

def Query(sql):
	db = pymysql.connect("localhost","root","stroops2020","CIN_Number",charset='utf8mb4')
	dbc = db.cursor(pymysql.cursors.DictCursor)
	output = {}
	qcnt = -1
	for statement in sqlparse.split(sql):
		qcnt = qcnt + 1
		dbc.execute(statement)
		db.commit()
		output[qcnt] = dbc.fetchall()
	if qcnt == 0:
		output = output[0]
	dbc.close()
	db.close()
	return output
f=open('comp.txt', 'r')
a=f.read()
f.close()
filepath="/var/www/seleniumtest/excel/acexcel.xlsx"
wb=openpyxl.load_workbook(filepath)
sheet=wb.active
rowcount = sheet.max_row
colcount = sheet.max_column
for i in  range(1,rowcount,1):
	if i>int(a):
		print('hiiiiiiiiiiii', str(i))
		we =sheet.cell(i,1).value
		print(we)
		try:
			Query('INSERT INTO CIN_NO(CIN) VALUES("'+str(we)+'")')
		except Exception as e:
			print(e)
	
		f=open('comp.txt', 'w+')
		f.write(str(i))
		f.close()

